#!/usr/bin/env python3

import os
import seaborn as sns
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from itertools import combinations
import socket
from difflib import SequenceMatcher
from copy import deepcopy
import re
from scipy import stats

from AHP import ahp_mat
import operator
from openpyxl import load_workbook
import pickle

from functools import reduce
## Helper function to get a product
def prod(iterable):
    return reduce(operator.mul, iterable, 1)
###

def getDataframe(filename):
    """
    Function to return the dataframe  
    """
    try:
        df_dict = pd.read_excel(filename, sheet_name=None, index_col=0)

        # This gets rid of the column name, which makes for an empty row. If we can live with the empty row, then we can use something else here
        for df in df_dict.values():
            # Remove the first and last 5 rows - boundary effects
            df = df.iloc[5:-5]
            df.index.name = None
        return df_dict
    except FileNotFoundError as e:
        raise e

def plot_host_df(df):
    """
    Function for plotting a single host df.
    Values in a host_df:
        "Time", "Duration", "Samples", "CPU Count", "Power",
        "CPU Load mean", "CPU Load max", "CPU Load std",
        "Used Memory mean", "Used Memory max", "Used Memory std",
        "Available Memory mean", "Available Memory min", "Available Memory std", 
        "Shared Memory mean", "Shared Memory std", "Shared Memory max",
        "Swap Available mean", "Swap Available std", "Swap Available min",
        "Swap Used mean", "Swap Used std", "Swap Used max"  
    """

    fig, axes = plt.subplots(3, 2, sharex=True)
    df_cpu = df.filter(regex='CPU L')
    df_sw_avail = df.filter(regex='Swap Av')
    df_mem_avail = df.filter(regex='Available Mem')
    df_mem_used = df.filter(regex='Used Mem')
    df_sw_used = df.filter(regex='Swap Use')
    df_mem_shared = df.filter(regex='Shared')
    used_cols = df_cpu.columns.values.tolist() + df_mem_avail.columns.values.tolist() + df_mem_used.columns.values.tolist() + df_sw_avail.columns.values.tolist() \
        + df_sw_used.columns.values.tolist() + df_mem_shared.columns.values.tolist()
    df_leftover = df.drop(labels=used_cols, axis=1)
    # df_memory = pd.concat([df_mem, df_swap], sort=False)
    ax1 = sns.lineplot(data=df_mem_avail, legend="full", ax=axes[0, 0])       #  dashes=False
    ax1.set(xlabel="Samples", ylabel="in MB")
    ax1.set_title('Available Memory')
    ax2 = sns.lineplot(data=df_mem_used, legend="full", ax=axes[0, 1])
    ax2.set(xlabel="Samples", ylabel="in MB")
    ax2.set_title('Used Memory')
    ax3 = sns.lineplot(data=df_cpu, legend="full", ax=axes[1, 0])
    ax3.set(xlabel="Samples", ylabel="in %")
    ax3.set_title('CPU Usage')
    ax4 = sns.lineplot(data=df_mem_shared, legend="full", ax=axes[1, 1])
    ax4.set(xlabel="Samples", ylabel="in MB")
    ax4.set_title('Shared Memory')
    ax5 = sns.lineplot(data=df_sw_used, legend="full", ax=axes[2, 0])
    ax5.set(xlabel="Samples", ylabel="in MB")
    ax5.set_title('Used Swap')
    ax6 = sns.lineplot(data=df_sw_avail, legend="full", ax=axes[2, 1])
    ax6.set(xlabel="Samples", ylabel="in MB")
    ax6.set_title('Available Swap')
    # sns.lineplot(data=df_leftover, legend="full", ax=axes[1, 1])
    print("Leftover values: {}".format(df_leftover.columns.values.tolist()))
    plt.show()

def summarize_node_df(df_dict):
    """
        receives a dictionary of node dataframes, which should be compared.
        Returns a dictionary of dataframes compiled by similar measures.
        Values in a node_df:
            "Time", "Duration", "Samples", "CPU Count",
            "Threads", "CPU Load mean", "CPU Load max", "CPU Load std",
            "PSS mean", "PSS std", "PSS max",
            "Swap Used mean", "Swap Used std", "Swap Used max",
            "Virtual Memory mean", "Virtual Memory std", "Virtual Memory max"
    """

    swap_df = pd.DataFrame()
    cpu_df = pd.DataFrame()
    virt_df = pd.DataFrame()
    pss_df = pd.DataFrame()
    leftover_df = pd.DataFrame()
  
    for node, df in df_dict.items():
        # Drop all std deviation values
        df.drop(list(df.filter(regex=" std")), axis=1, inplace=True)
        _sw = df.filter(regex='Swap')
        _c = df.filter(regex='CPU L')
        _vir = df.filter(regex='Virtual')
        _pss = df.filter(regex='PSS')
        used_cols = _sw.columns.values.tolist() + _c.columns.values.tolist() + _vir.columns.values.tolist() + _pss.columns.values.tolist() 
        _lefto  = df.drop(labels=used_cols, axis=1)
        dfs = [_sw, _c, _vir, _pss, _lefto]
        
        # turn the names unique with the node dictionary names
        # clean the indices
        for _df in dfs:
            new = node.split('_')[1]+'_'
            colnames = [new+name for name in _df.columns.values.tolist()]
            _df.columns = colnames
            _df.index = np.round(_df.index.to_series().to_numpy(), decimals=1)
        
        swap_df = pd.concat([swap_df, _sw], axis=1, sort=False)
        cpu_df = pd.concat([cpu_df, _c], axis=1, sort=False)
        virt_df = pd.concat([virt_df, _vir], axis=1, sort=False)
        pss_df = pd.concat([pss_df, _pss], axis=1, sort=False)
        leftover_df = pd.concat([leftover_df, _lefto], axis=1, sort=False)

    print("Test")
    compiled_df = {}
    compiled_df["Swap"] = swap_df
    compiled_df["CPU"] = cpu_df
    compiled_df["Virt"] = virt_df
    compiled_df["Pss"] = pss_df
    compiled_df["Leftovers"] = leftover_df
    return compiled_df
    
def plot_node_df_dict(df_dict, plot_name=""):
    """
     A function which takes a dictionary sorted by nodes and compiles it 
    """

    elems = int(np.ceil(np.sqrt(len(df_dict))))
    # fig = plt.figure()
    fig, axes = plt.subplots(elems, elems, sharex=True)
    axes = axes.reshape(-1)
    for i, (name, df) in enumerate(df_dict.items()):
        sns.lineplot(data=df, legend="full", dashes=False, ax=axes[i])
        axes[i].set(xlabel="Samples")
        axes[i].set_title(name)
        axes[i].set_ylim(0,)
    plt.suptitle(plot_name)
    plt.show()
    print("Test Done")


def compare_dicts(big_dict, matching="hosts"):
    """
        Function to compare a dictionary of  dictionaries for the same information in them.
        Mostly to be used with node dictionaries.
        For now only for **2** dictionary comparison
    """
    
    node_dictionaries = []
    if matching is "hosts":
        print("Matching by Hostnames")
    elif matching is "filename":
        print("Matching by filenames")
    else:
        print("No suitable matching specified. Matching By Hostnames")
        matching="hosts"

    for pair in combinations(big_dict.items(), 2):
        # pair is a key, value tuple
        node_dictionaries.append(compare_two_dicts(pair[0], pair[1], matching=matching))
    
    # Consolidating the dictionary combination
    consolid_dict = {}
    consolid_dict = deepcopy(node_dictionaries[0])
    for bdicts in node_dictionaries[1:]:
        for node, small_dicts in bdicts.items():
            # if the node is already in there
            if node in consolid_dict.keys():
                for key in small_dicts.keys():
                    if key not in consolid_dict[node].keys():
                        new_dict = consolid_dict[node]
                        new_dict[key] = small_dicts[key]
                        consolid_dict[node] = new_dict
            else:
                consolid_dict[node] = small_dicts
    return consolid_dict



def compare_two_dicts(dict_1, dict_2, matching="hosts"):
    """
     Granular function comparing the keys for two dictionaries to see if they match.
     Is called by `compare_dicts` function
    """
    # dict_1 and _2 being passed in are key, value tuples - have to be indexed
    process_dict = {}
    for hn_name in dict_1[1].keys():
        namesplit = hn_name.split('_')
        h_name1 = ip_lookup(namesplit[0].lower())
        nname1 = '_'.join(namesplit[1:]).lower()
        for key in dict_2[1].keys():
            namesplit2 = key.split('_')
            h_name2 = ip_lookup(namesplit2[0].lower())
            nname2 = '_'.join(namesplit2[1:]).lower()

            # Double-sided matching process for nodes
            if (nname1 in nname2) or (nname2 in nname1):
                # reorder matching. Sort the dfs according to NODES
                # Get the DFs
                combined_dict = {}
                df_1 = dict_1[1][hn_name]
                df_2 = dict_2[1][key]
                if matching is "hosts":
                    combined_dict[h_name1] = df_1
                    combined_dict[h_name2] = df_2
                elif matching is "filename":
                    pre1 = dict_1[0].split('_')[:-1]
                    pre1.append(h_name1)
                    k1 = '_'.join(pre1)
                    pre2 = dict_2[0].split('_')[:-1]
                    pre2.append(h_name2)
                    k2 = '_'.join(pre2)
                    combined_dict[k1] = df_1
                    combined_dict[k2] = df_2
                else:
                    combined_dict[h_name1] = df_1
                    combined_dict[h_name2] = df_2
                # turn into the big dictionary entry
                nname = get_overlap(nname1, nname2)
                print("Match found: {} and {}, Consolidated into: {}".format(nname1, nname2, nname))
                process_dict[nname] = combined_dict
                break
        # [print("Match found: {}".format(nname)) for key in dict_2[1].keys() if nname in key.lower()]


    if len(process_dict) < 1:
        process_dict = None
    return process_dict

def compare_host_dicts(host_dict):
    """
        Function to compare two host dictionaries
    """
    big_dict = {}
    for c_name, h_dict in host_dict.items():
        cnamesplit = c_name.split('_')
        cname = '_'.join(cnamesplit[:-1])
        for hname, df in h_dict.items():
            hname = ip_lookup(hname)
            new_name = '_'.join([cname, hname])
            big_dict[new_name] = df


    return big_dict

def drizzleUpDictionary(host_dict, device, out_dict):
    """
        Function to drizzle up the massive dictionaries.
        Case order in key as a tuple:
        1. Device
        2. partial/full
        3. Compressed/Uncompressed
        4. Wifi/Ethernet
        
    """
    # out_dict = {}
    for cname, hdict in host_dict.items():
        cnamesplit = cname.split('_')
        if cnamesplit[0].lower() == "wifi":
            case = tuple([device, "partial", "compressed", "wifi"])
        elif cnamesplit[0].lower() == "full":
            case = tuple([device, "full", "compressed", "wifi"])
        elif cnamesplit[0] == "wireUncompr":
            case = tuple([device, "partial", "uncompressed", "ethernet"])
        elif cnamesplit[0] == "wireCompr":
            case = tuple([device, "partial", "compressed", "ethernet"])
        elif cnamesplit[0] == "wireFullCompr":
            case = tuple([device, "full", "compressed", "ethernet"])
        elif cnamesplit[0] == "pc":                                   # Do not need this case for the host comparison
            case = tuple([device, None, "compressed", None])
        elif cnamesplit[0] == "compr": 
            case = tuple([device, "partial", "compressed", "wifi"])
        else:
            raise TypeError("Name not known: {}".format(cnamesplit[0]))

        # Get the device host dictionary
        for k, hdf in hdict.items():
                if device in k:
                    df = hdf
                    break
        
        # Try to append to the list of dfs. If the list does not exist (Key Error), create it
        try:
            caselist = out_dict[case]
            caselist.append(df)
            out_dict[case] = caselist
        except KeyError:
            out_dict[case] = [df]
        # caselist = out_dict.setdefault(case, [df])
        # if len(caselist) > 1:
        #     caselist.append(df)
        # out_dict[case] = caselist
    # return out_dict

def removeNanRows(host_dict):
    """
        Function to remove nan rows from all the dfs
    """

    for case, df_dict in host_dict.items():
        for param, df in df_dict.items():
            df.dropna(inplace=True)


def filepaths(directory_string, file_string=None):
    """
        Searches from the base directory of the file for the directory with the directory string
        and in that directory searches for the file string.
        Returns:: [str] list of full filepaths.
    """
    file_dicts = {}
    dirpath = os.path.join(os.path.abspath(os.path.dirname(__file__)), '..')
    for path in Path(dirpath).rglob('*'+directory_string+'*'):
        if os.path.isdir(path):
            for fname in Path(path).rglob('*'+file_string+'*'):
                try:
                    f = os.path.splitext(os.path.basename(fname))[0]
                    file_dicts[f] = getDataframe(fname)
                    print("Found Dataframes for file: {} in folder {}".format(f, os.path.basename(os.path.dirname(fname))))
                except FileNotFoundError:
                    print("Could not open file {}. Continuing".format(fname))
    
    # For error catching in outer function
    if len(file_dicts) < 1:
        file_dicts = None
    return file_dicts
    

def plot_processes(big_dict, filter_list=["Swap", "CPU L", "Virtual", "PSS", "Threads"]):
    """
        plotting the same processes (Nodes) if and when they are running on different plattforms
        Requires a list of dictionaries (lods) in the following format: 
        Every item in the list is a dictionary, which holds node names. For each of these node names, there is a separate dictionary, with the key
        being the identificator of the host and the value being the dataframe with the values
    """
    node_dict = {}
    for node, ch_dict in big_dict.items():
            
            # host_dict is itself a dictionary, where the keys are the hosts and the values are the dfs
            node_dict[node] = process_resources(ch_dict, filter_list)
            print("The above values are for node: {}".format(node))
            print("===============================")

    for node, df_dict in node_dict.items():
        plot_node_df_dict(df_dict, node)

    print("Test")
    # node_dict now holds dfs sorted by: dict[node] = dict[filters] = 
            


def process_resources(dictionary, filter_list, leftover="leftover"):
    """
        A function, which takes in a dictionary with a name and a long df,
        and returns a dictionary of dfs
        Params: Filter list, with default values: Swap, CPU L, Virtual and PSS  
    """
    df_dict = {}
    for i in range(len(filter_list)+1):
        if i == len(filter_list):
            df_dict[leftover] = pd.DataFrame()
        else:
            df_dict[filter_list[i]] = pd.DataFrame()
  
    for name, df in dictionary.items():
        # Drop all std deviation values
        df.reset_index(drop=True, inplace=True)       # has to be done, because the time is not aligned across the hosts
        df.drop(list(df.filter(regex=" std")), axis=1, inplace=True)
        # used_cols = []
        df.index = np.round(df.index.to_series().to_numpy(), decimals=1)
        for fil in filter_list:
            _tmp = df.filter(regex=fil)
            df.drop(_tmp.columns.values.tolist(), axis=1, inplace=True)
            # used_cols = used_cols + _tmp.columns.values.tolist()
            new_colnames = [name+'_'+col for col in _tmp.columns.values.tolist()]
            _tmp.columns = new_colnames
            df_dict[fil] = pd.concat([df_dict[fil], _tmp], axis=1, sort=False)
            # df.drop(_tmp.columns)
        new_colnames = [name+'_'+col for col in df.columns.values.tolist()]
        df.columns = new_colnames
        df_dict[leftover] = pd.concat([df_dict[leftover], df], axis=1, sort=False)
    
    for key, value in df_dict.items():
        print("Df for: {}".format(key))
        print(value.head())

    return df_dict

def compare_vals(bdict, filter_list=["^(CPU L).*(max)$", "^(Used).*(Mem).*(max)$", "^(Avail).*(Mem).*(min)$", "^(Swap).*(U).*(max)$"]):
    """ 
    Function to take in the big dictionary and compare values of interest.
    At Host level
    Values to compare
        * CPU Load Max
        * Phymem used Max
        * Phymem avail Min

    Hard Checks:
        * Swap Used max > 0?
    """
    # Prefilter the list 
    # Get the first df
    df = list(bdict.values())[0]
    collist = df.columns.values.tolist()
    
    fillist = filtercolumns(collist, filter_list)
    outdfdict = {}
    for fil in fillist:
        outdfdict[fil] = pd.DataFrame()
    
    # With the list of headers, filter the dfs
    for name, df in bdict.items():
        ndf = df.reset_index(drop=True)
        # ndf.index = np.round(df.index.to_series().to_numpy(), decimals=1)
        ndf = ndf.filter(items=fillist)
        
        for col in ndf.columns.values.tolist():
            fdf = ndf.filter(like=col, axis=1)
            fdf.columns = [name]
            outdfdict[col] = pd.concat([outdfdict[col], fdf], axis=1, sort=False)
            
    print("test Done")
    return outdfdict

def newcomparisonfunc(bdict, filter_list=["^(Samples)","^(CPU L).*(max)$", "^(Used).*(Mem).*(max)$", "^(Avail).*(Mem).*(min)$", "^(Swap).*(U).*(max)$", "^(Pow).*"]):
    """
        Function to compare the comparisons
        has to put out a dictionary of dictionaries
    """

    df = list(bdict.values())[0][0]
    collist = df.columns.values.tolist()

    fillist = filtercolumns(collist, filter_list)
    outdfdict = {}
    for k in bdict.keys():
        ndict = {}
        for fil in fillist:
            df = pd.DataFrame()
            ndict[fil] = df
        outdfdict[k] = ndict

    for key, listofdfs in bdict.items():
        for i, df in enumerate(listofdfs): 
            ndf = df.reset_index(drop=True)
            ndf = ndf.filter(items=fillist)

            for col in ndf.columns.values.tolist():
                fdf = ndf.filter(like=col, axis=1)
                fdf.columns = [i]
                prior_df = outdfdict[key][col]
                post_df = pd.concat([prior_df, fdf], axis=1, sort=False)
                outdfdict[key][col] = post_df

    print("Internal Test done")
    return outdfdict



def consolid_values(bdict, skipna=True, alpha=1e-3, perc=0.9):
    """
        A function to turn the big dictionary into a dataframe by using consolidated values
    """
    # cpu_df = bdict['CPU Load max']
    # normality = normtest(cpu_df.values)
    # print("Normality test failed: {} of {}. Using Quantiles".format(normality[np.where(normality<alpha)].shape[0],normality.shape[0]))
    # print("Normality test Done. Proceeding with array building")
    # print("Looking for max values")
    indices = list(bdict.keys())
    val = list(bdict.keys())[0]
    case = bdict[val]
    columns = list(case.keys())
    # columns = list(list(bdict.keys())[0].values().keys())
    bdf = pd.DataFrame(data=None, index=indices, columns=columns)
    for case, dfdict in bdict.items():
        for key, df in dfdict.items():
            prevals = df.to_numpy()
            prevals = prevals[np.nonzero(prevals)]
            val = np.nanquantile(prevals,perc)
            bdf.at[case,key] = val
        # bdf.at[key] = df.max(skipna=skipna)
        # prevals = df.values.to_numpy()
        # prevals = prevals[np.nonzero(prevals)]
        # vals = np.nanquantile(prevals, perc, axis=0)
        # bdf.at[key] = vals

    # print(bdf.head())
    # print("Big Df done")
    return bdf

def getfaults(bdict, target_list):
    """
        Function to count the faults of a system.
        Calculated over the CPU load mean
        appends it to the target-df
    """
    ser = pd.Series(index=target_list, name="Faults")
    for name in target_list:
        df = bdict[name]
        ndf = df.reset_index(drop=True, inplace=False)
        ndf = ndf.filter(regex="^(CPU L).*(mean)$")
        ct = np.count_nonzero(ndf.values < 0.1)
        ser.at[name] = ct
    return ser

def redeffaults(bdict, thresh=7.5):
    """
        Fault redefinition: Adds Boolean DF for each case. Removes the "Samples" Entry
    """
    for dfdict in bdict.values():
        df = dfdict["Samples"]
        mask = df < thresh
        dfdict["Faults"] = deepcopy(mask)
        del dfdict["Samples"]

def maskfaults(bdict):
    """
        Masking the faults
    """
    faults_ct = {}
    for k, dfdict in bdict.items():
        maskdf = dfdict["Faults"]
        mask = maskdf.any(axis=1)
        faults_ct[k] = mask.sum()
        for _, df in dfdict.items():
            df = df[~mask]
        del dfdict["Faults"]

    return faults_ct

def getpower(bdict, pinom=4.1):
    """
        Function to calculate the median power usage.
        receives a list
        returns a Series
    """

    ser = {}
    for k, v in bdict.items():
        df = v["Power"]
        name = k[0]
        ndf = df.reset_index(drop=True, inplace=False)
        # ndf = ndf.filter(like="Power")
        if np.sum(ndf.values) < 0: # Since the values if the file is not found is negative, we use the CPU_Load_max
            alt = v["CPU Load max"]
            if "pi" in name:
                # vol = pivol
                # amp = piamp
                # watt = vol*amp
                # power = np.mean((10*watt*alt.values))
                # This gives waaay to high values, use nominal wattage instead: 4.1
                avg = np.mean(alt.values)
                power = avg * pinom * 10
            elif "sef" in name:
                power=-1.0
            else:
                print("Do not know this kind of device")
                power=-1.0
        else:
            power = np.mean(ndf.values)
        ser[k] = power
    return ser


def normtest(arr, alpha=5e-2, axis=0, nan_policy='omit'):
    """
        A function to test for normality - to use with the CPU Load. Returns true if the 0-Hypothesis can be rejected
        (Commonly: 0-Hypothesis: it comes from a normal distribution)
    """
    _, p = stats.normaltest(arr, axis=0, nan_policy=nan_policy)
    if p.shape[0] > 1:
        return p
    else:
        if p < alpha:
            return False
        else:
            return True
        

def filtercolumns(collist, filters=["^(CPU L).*(max)$", "^(Used M).*(max)$", "^(Avail).*(Mem).*(min)$", "^(Swap U).*(max)$"]):
    """
        Test function to find out the filters for the columns
    """

    outputl = []
    for fil in filters:
        r = re.compile(str(fil))
        outputl+=list(filter(r.search, collist))
    return outputl
    
def ip_lookup(ip):
    """
        Helper function. Looks up whether a hostname is an IP or a hostname and ALWAYS returns the hostname
    """
    if len(ip.split('.')) > 2:
        ### BAD: hardcoded lookup here:
        if ("1.120" in ip) or ("1.132" in ip):
            return "nico-nano"
        try:
            hostname = socket.gethostbyaddr(ip)
            return hostname
        except:
            return ip
    else:
        return ip

def get_overlap(s1, s2):
    """
        Helper function to get the biggest overlapping substring between the strings, from [here](https://stackoverflow.com/questions/14128763/how-to-find-the-overlap-between-2-sequences-and-return-it)
        Used to find the node name overlap
    """
    s = SequenceMatcher(None, s1, s2)
    pos_a, pos_b, size = s.find_longest_match(0, len(s1), 0, len(s2))
    return s1[pos_a:pos_a+size]

def process_faults(vec):
    """
        Function to return an array of the relative importance of faults. Uses the natural logarithm to scale accordingly
    """
    def calclns(vec):
        """
            Function to calculate the natural logarithms and return them
        """

        lnvec = np.zeros_like(vec)
        idxs = np.where(vec>=1.0)
        lnvec[idxs] = np.log(vec[idxs])
        return lnvec

    def lncomparison(lnval1, lnval2):
        """
            output array
            indices
            lnvec
        """
        # piecewise comparison
        if lnval1 == lnval2:
            return 1.0
        elif lnval1 < 0.9:
            return (9.0-lnval2)
        elif lnval2 < 0.9:
            return 1.0/(9.0-lnval1)
        else:
            out = 1.0+(max([lnval1, lnval2]) - min([lnval1, lnval2]))
            if lnval1 > lnval2:
                return 1.0/out
            else:
                return out

    lnvec = calclns(vec)
    outputarr=np.ones((lnvec.size, lnvec.size))
    maxln = np.max(lnvec)
    for i in range(lnvec.size):
        for j in range(lnvec.size):
            if i>=j:
                continue
            else:
                outputarr[i,j] = lncomparison(lnvec[i], lnvec[j])
    return outputarr

def flts_naive(vec):
    """
        Naive way of putting the vector in: if one is > 0, assign that a 9
    """
    
    outarr = np.ones((vec.size, vec.size))
    for i in range(vec.size):
        for j in range(vec.size):
            if i >=j:
                continue
            elif vec[i] == vec[j]:
                outarr[i,j] = 1.0
            elif vec[j] < 1.1:
                if vec[i] < 1.1:
                    outarr[i,j] = 1.0
                else:
                    outarr[i,j] = 1.0/9.0            
            elif vec[i] < 1.1:
                outarr[i,j] = 9.0
            else:
                outarr[i,j] = np.log10(vec[j])/np.log10(vec[i])

    return outarr


# Comparing Power Things
def powerfromVI(curr, vol=5.0):
    """
        Function to calculate the wattage from voltage and current
    """
    return curr*vol

def powerfromNom(vol=None, amph=4.0,t0=0.33, volpcell=3.7, cells=3, safety=0.8):
    """
        calculating P from values common to UAVs. Flight time at hover, battery cells, battery voltage, nominal ampere hour capacity rating
    """
    if vol is None:
        vol = volpcell*cells
    p0 = (vol * amph * safety) / t0
    return p0
        
def flighttime(pe, p0, c):
    """
        calculating the new flight time using the additional power draw
    """
    tf = (c) / (p0 + pe)
    return tf

def powerstuff(vec, safety=0.8 , volpcell=3.7, cells=3, amph=4.0, t0=0.333):
    """
        Calculating the ballpark power draw that we are using for the uav
        params = safety capacity, nominal voltage, amperehour-rating, flight time at hover
        Calculating the relative reduce in flighttime - using hover values and reasonable assumptions
    """
    tf_vec = np.zeros_like(vec)
    pe_vec = np.zeros_like(vec)
    vec = vec / 1000.0            # to get amperes
    p0 = powerfromNom(amph=amph,t0=t0,safety=safety,cells=cells,volpcell=volpcell)
    powfunc = np.vectorize(powerfromVI)
    pe_vec = powfunc(vec)
    c = volpcell*cells*amph*safety
    tffunc = np.vectorize(flighttime)
    tf_vec = tffunc(pe_vec,p0=p0,c=c)
    diffvec = (t0-tf_vec)/t0
    arr = np.ones((vec.size, vec.size))
    for i in range(vec.size):
        for j in range(vec.size):
            if i>=j:
                continue
            else:
                arr[i,j] = 1.0/comparestraight(diffvec[i], diffvec[j])

    return arr

# Comparing Weight and Volume 
def weightcalc(w1, w2, w0=1382.0):
    """
        Just use the relative weights, with hard limits @ 9 and 1/9
    """
    w1rel = w1 / (w1+w0)
    w2rel = w2 / (w2+w0)
    relw = w1rel/w2rel
    if relw > 9.0:
        return 9.0
    elif relw < (1.0/9.0):
        return 1.0/9.0
    else:
        return relw

def weightstuff(df, dictnom):
    """
        function to calculate the relative weight
    """
    names = df.index.values.tolist()
    outarr = np.ones((len(names), len(names)))
    w0 = dictnom["full"]
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i >= j:
                continue
            else:
                if "pi" in names[i][0]:
                    w1 = dictnom["pi"]
                elif "nano" in names[i][0]:
                    w1 = dictnom["nano"]
                else:
                    raise ValueError("No known volume")
                if "pi" in names[j][0]:
                    w2 = dictnom["pi"]
                elif "nano" in names[j][0]:
                    w2 = dictnom["nano"]
                else:
                    raise ValueError("No known volume")
                outarr[i,j] = 1.0/weightcalc(w1,w2,w0)
    return outarr
    

def volstuff(df, dictnom):
    """
        using the dataframe columns and the dictionary to construct a relative array
    """
    names = df.index.values.tolist()
    outarr = np.ones((len(names), len(names)))
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i >= j:
                continue
            else:
                if "pi" in names[i][0]:
                    v1 = dictnom["pi"]
                elif "nano" in names[i][0]:
                    v1 = dictnom["nano"]
                else:
                    raise ValueError("No known volume")
                if "pi" in names[j][0]:
                    v2 = dictnom["pi"]
                elif "nano" in names[j][0]:
                    v2 = dictnom["nano"]
                else:
                    raise ValueError("No known volume")
                outarr[i,j] = volcalc(v1,v2)
    return outarr

def volcalcalt(v1,v2):
    """
        Alternative volume calculation. Using the product of all values
    """
    vol1=prod(v1)
    vol2=prod(v2)
    return comparestraight(vol1,vol2)


def volcalc(vec_1, vec_2):
    """
        Taking in 2 vectors of l x w x h values.
        Calculates the highest divide by the lowest value of each vector
    """

    v1 = volratio(vec_1)
    v2 = volratio(vec_2)
    return comparestraight(v1,v2)
    
def volratio(vec):
    """
        Function to calculate the ratio of lowest side to longest side
    """
    
    minval = np.min(vec)
    maxval = np.max(vec)
    return maxval/minval
  
def powervals(val1, val2, fact):
    """
        Function to calculate the preferred power values
    """
    
    if val1 == val2:
        return 1.0
    else:
        res = (val1/val2)
        if res<1.0:
            return fact/res
        else:
            return 1/(fact*res)

# Comparing directly
# Use for: CPU usage and CPU free
# Also for % of memory used - first divide by nominal value
# Also for free memory - use ln 
def comparestraight(val1, val2,scale=1.0):
    """
        Function for direct comparison. if values over 9, set to 9    
    """
    
    rel = scale*(val1/val2)
    if rel > 9.0:
        return 9.0
    elif rel < (1.0/9.0):
        return (1.0/9.0)
    else:
        return rel

def cpu_usage(vec):
    """
        Function to return comparable cpu usage values as an array/ uses comparestraight
    """

    outarr = np.ones((vec.size, vec.size))
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i >= j:
                continue
            else:
                outarr[i,j] = comparestraight(vec[i], vec[j])
    
    return outarr


def cpu_free(vec, scale=1.0):
    """
        Function to return compared cpu free values as an array. uses comparestraight
    """
    vec = 1.0 - (vec/100)
    outarr = np.ones((vec.size, vec.size))
    # lnvec = np.log(vec)
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i >= j:
                continue
            else:
                outarr[i,j] = comparestraight(vec[i], vec[j], scale=scale)
    # print(outarr)
    return outarr


def memrel(val, nom):
    """
        Function to return the % of memory used, using the nominal value        
    """
    return val/nom

def freemem(vec):
    """
        Function to return the free memory. Uses Comparestraight. Does not use the ln
    """
    lnvec = np.log(vec)
    outarr = np.ones((vec.size, vec.size))
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i >= j:
                continue
            else:
                outarr[i,j] = comparestraight(lnvec[i], lnvec[j])
    
    return outarr

def memperc(df, dictnom, column_name):
    """
        Alternative method to calculate the relative free memory using the nominal values of the devices
    """

    ser = df[column_name]
    vec = np.zeros_like(ser.values)
    # cases = df.index.values.tolist()
    for i, case in enumerate(df.index.values.tolist()):
        val = df.loc[case, column_name]
        if "pi" in case[0]:
            nomm = dictnom["pi"]
        elif "nano" in case[0]:
            nomm = dictnom["nano"]
        else:
            raise TypeError("Unknown Nominal memory value for this type of device")
        vec[i] = memrel(val,nomm)
    outarr = np.ones((vec.size, vec.size))
    for i in range(outarr.shape[0]):
        for j in range(outarr.shape[1]):
            if i>=j:
                continue
            else:
                outarr[i,j] = comparestraight(vec[i], vec[j])
    return outarr


def makeMultiIndex(df, levelnames = ("device", "load", "compression", "network")):
    """
        Method to turn the tuples into a multiindex
    """
    tuples = df.index.values.tolist()
    columns = df.columns.values.tolist()
    data = deepcopy(df.to_numpy())
    multiindex = pd.MultiIndex.from_tuples(tuples,names=levelnames)
    newdf = pd.DataFrame(data=data, index=multiindex, columns=columns)
    return newdf

def parse_file(fname):
    """
        Function to parse a xlsx file
    """
    out_dict={}
    xls = pd.ExcelFile(fname)
    dfL1 = pd.read_excel(xls, "L1", index_col=0)
    dfL1.fillna(1.0, inplace=True)
    out_dict["L1"] = dfL1
    dfL2 = pd.read_excel(xls, "L2", index_col=0)
    dfL2.fillna(1.0, inplace=True)
    out_dict["L2"] = dfL2
    dfL31 = pd.read_excel(xls, "L3.1", index_col=0)
    dfL31.fillna(1.0, inplace=True)
    out_dict["L3.1"] = dfL31
    dfL32 = pd.read_excel(xls, "L3.2", index_col=0)
    dfL32.fillna(1.0, inplace=True)
    out_dict["L3.2"] = dfL32
    return out_dict

def writeToFile(df, fname, sheet_name="Results"):
    """
        File to write the results into the same document as it is being read from
    """
    wb = load_workbook(fname)
    if sheet_name in wb.sheetnames:
        return False
    else:
        writer = pd.ExcelWriter(fname, engine = 'openpyxl')
        writer.book = wb
        df.to_excel(writer, sheet_name=sheet_name)
        writer.save()
        writer.close()
        return True
        
def nodestuff(pdir):
    """
        The stuff for reading through the node files
    """
    bdict = {}
    filetype='node'
    filedicts = filepaths("results_nano", filetype)
    pc_dicts = filepaths("results_pc", filetype)
    rasp_dicts = filepaths("results_pi", filetype)

    drizzleUpDictionary(filedicts,"nano",bdict)
    drizzleUpDictionary(pc_dicts,"sef",bdict)
    drizzleUpDictionary(rasp_dicts,"pi",bdict)

    print("Test Done")


if __name__=="__main__":

    sns.set()

    # 1. get the File which results we are interested in
    counter = 1
    filetype = 'node'
    result_host = '_pc'
    # result_host = '_nano'
    parentDir = os.path.dirname(__file__)
    fname = os.path.abspath(os.path.join(parentDir, '..','results'+result_host,'case_'+str(counter)+result_host+'_'+filetype+'.xlsx'))
    devices = ["nano", "sef", "pi"]
    compression = ["compressed", "uncompr"]
    connection = ["WiFi", "Ethernet"]
    composition = ["Full", "Partial"]
    
    # TODO: NODES
    nodestuff(parentDir)

    # print(first_df.columns)
    pdir = os.path.abspath(os.path.join(parentDir,'..','tmp'))
    dfpickle = 'ConsolidDF.pickle'
    dfpf = os.path.join(pdir,dfpickle)
    try:
        with open(dfpf, 'rb') as f:
            cdf = pickle.load(f)
        print("File {} loaded successfully".format(dfpf))
    except FileNotFoundError:
        print("File {} Not found. Proceeding with loading prior data and writing file".format(dfpf))    
        relpfname = 'OutDict.pickle'

        outpickf = os.path.join(pdir, relpfname)
        try:
            with open(outpickf, 'rb') as f:
                out_dict = pickle.load(f)
            print("Out_df successfully loaded")
        except FileNotFoundError:
            filetype = 'host'
            filedicts = filepaths("results_nano", filetype)
            pc_dicts = filepaths("results_pc", filetype)
            rasp_dicts = filepaths("results_pi", filetype)
            out_dict = {}
            drizzleUpDictionary(filedicts, "nano", out_dict)
            drizzleUpDictionary(pc_dicts, "sef", out_dict)
            drizzleUpDictionary(rasp_dicts, "pi", out_dict)

            with open(outpickf, 'wb') as f:
                pickle.dump(out_dict, f)
            print("Pickle File dumped")
    
        # TODO: CONTINUE HERE WITH THE CONSOLIDATION OF THE VALUES
        # Filterlist = ["^(CPU L).*(max)$", "^(Used M).*(max)$", "^(Avail).*(Mem).*(min)$", "^(Swap U).*(max)$"]
        case_dict = newcomparisonfunc(out_dict)
        # 3. Remove all rows that contain NaN values
        removeNanRows(case_dict)

        # 4. redefining faults - for each case, return a tuple where the case is true
        redeffaults(case_dict)

        # 5. Using the faults to mask the results
        faults_ct = maskfaults(case_dict)

        # 6. Use the out_dict to get the power values
        pow_dict = getpower(case_dict)

        # Extract the values of interest
        cdf = consolid_values(case_dict, perc=0.75)
        for k,v in faults_ct.items():
            cdf.at[k, "Faults"] = v
        for k,v in pow_dict.items():
            cdf.at[k, "Power"] = v
        cdf.fillna(value=0,inplace=True)
        # print(cdf)
        with open(dfpf, 'wb') as f:
                pickle.dump(cdf, f)
        print("File {} dumped successfully".format(dfpf))

    pivol = np.array([0.07, 0.016, 0.06])
    nanovol = np.array([0.12, 0.062, 0.06])

    dictvol = {}
    dictvol["pi"] = pivol
    dictvol["nano"] = nanovol

    dictwt = {}
    dictwt["pi"] = 0.044
    dictwt["nano"] = 0.252
    dictwt["full"] = 1.322

    dictmem = {}
    dictmem["nano"] = 4096.0
    dictmem["pi"] = 512.0

    # Get the FULL DF 
    vol_of_pi = prod(pivol)*1000
    vol_of_nano = prod(nanovol)*1000

    # Removing the SITL stuff
    for keys in cdf.index.tolist():
        if "sef" in keys[0]:
            key = keys
            continue
        elif "pi" in keys[0]:
            cdf.at[keys,"Volume"] = vol_of_pi
            cdf.at[keys, "Weight"] = dictwt["pi"]
        elif "nano" in keys[0]:
            cdf.at[keys,"Volume"] = vol_of_nano
            cdf.at[keys, "Weight"] = dictwt["nano"]
    # pc_ser = cdf[("sef", None, "compressed", None)]
    sth = cdf.T
    sth_new = deepcopy(sth[key])
    # pc_ser = deepcopy(cdf.loc[key])
    cdf.drop(key, axis=0, inplace=True)
    # print(cdf)
    # print(cdf)
    newdf = makeMultiIndex(cdf)
    # print(newdf)
    xlf = os.path.join(pdir,'MultiIndex.xlsx')
    if not os.path.exists(xlf):
        with pd.ExcelWriter(xlf) as writer:
            newdf.to_excel(writer,merge_cells=False)
        print("File {} Did not exist. Written".format(xlf))
    else:
        print("File {} Already exists. Not written.".format(xlf))

    sdf = newdf.xs('wifi', level='network')
    sxlf = os.path.join(pdir, 'FinalOptions.xlsx')
    if not os.path.exists(sxlf):
        with pd.ExcelWriter(sxlf) as writer:
            sdf.to_excel(writer)
        print("File {} Did not exist. Written".format(sxlf))
    else:
        print("File {} Already existed. Not Written".format(sxlf))

    print(sdf)
    # Getting the dataframes out - using the values from the original df
    # memfreearr = freemem(reddf.loc["Available Memory min"].to_numpy().astype(np.float64))
    memfreearr = memperc(sdf, dictmem, column_name="Available Memory min")
    memfreedf = pd.DataFrame(data=memfreearr, index=sdf.index, columns=sdf.index)
    cpufreearr = cpu_free(sdf["CPU Load max"].to_numpy().astype(np.float64))
    cpufreedf = pd.DataFrame(data=cpufreearr, index=sdf.index, columns=sdf.index)
    cpuusedarr = cpu_usage(sdf["CPU Load max"].to_numpy().astype(np.float64))
    cpuuseddf = pd.DataFrame(data=cpuusedarr, index=sdf.index, columns=sdf.index)
    # fltsarr = process_faults(cdf.loc["Faults"].to_numpy().astype(np.float64))
    fltsarr = flts_naive(sdf["Faults"].to_numpy().astype(np.float64))
    fltsdf = pd.DataFrame(data=fltsarr, columns=sdf.index, index=sdf.index)
    pwrarr = powerstuff(sdf["Power"].to_numpy().astype(np.float64))
    pwrdf = pd.DataFrame(data=pwrarr, columns=sdf.index, index=sdf.index)
    memusedarr = memperc(sdf, dictmem, column_name="Used Memory max")
    memuseddf = pd.DataFrame(data=memusedarr, index=sdf.index, columns=sdf.index)
    wtarr = weightstuff(sdf, dictwt)
    wtdf = pd.DataFrame(data=wtarr, index=sdf.index, columns=sdf.index)
    volarr = volstuff(sdf, dictvol)
    voldf = pd.DataFrame(data=volarr, index=sdf.index, columns=sdf.index)
    # print(memuseddf)
    # print(wtdf)
    # print(voldf)
    # print(fltsdf)
    print("=====================================================================")
    print("First Phase done. Host results are collected. Proceeding with the AHP.")
    print("=====================================================================")

    # Putting the stuff into AHPs
    # 1: Putting it into a dictionary
    df_dict = {}
    df_dict["CPU Used"] = cpuuseddf
    df_dict["Mem Used"] = memuseddf
    df_dict["CPU Free"] = cpufreedf
    df_dict["Mem Free"] = memfreedf
    df_dict["Power"] = pwrdf
    df_dict["Weight"] = wtdf
    df_dict["Size"] = voldf
    df_dict["Faults"] = fltsdf

    ahp_dict = {}
    for key, df in df_dict.items():
        ahp = ahp_mat(df,name=key)
        # print(ahp.df)
        # print(ahp.df)
        if ahp.getconsistency():
            print("{} is consistent. Adding to dictionary".format(key))
            ahp_dict[key] = ahp
        else:
            print("{} Not consistent. Ratio: {}".format(key,ahp.consratio))
    print("Required consistency: {}".format(ahp.CONSISTENT))
    
    # Section on Pickling. RUN ONLY ONCE
    picklefile = os.path.join(pdir, 'RelDict.pickle')
    if not os.path.exists(picklefile):
        print("Picklefile {} does not exist. Dumping.".format(picklefile))
        with open(picklefile, 'wb') as f:
            pickle.dump(ahp_dict, f)
    else:
        print("File {} already exists. Not Writing".format(picklefile))

        

    # ALL of them are consistent. Now do our own relative weighting
    l1names = ["Weight", "Size", "Power", "Computation"]
    l2names = ["Performance", "Compatibility","Reliability"]
    l3names = ["CPU", "Memory"]

    #### Parse the Excel File
    fname = "Case4.xlsx"
    floc = os.path.join(pdir,fname)
    weightsdict = parse_file(floc)

    ahp_wts_dict = {}
    for k, ldf in weightsdict.items():
        ahp = ahp_mat(ldf, name=k)
        print("Level: {}, Consistency, {}, Eigenvalues:\n{}".format(
            k, ahp.consratio, ahp.eigdf
        ))
        if not ahp.getconsistency():
            print("Level {} Not consistent. Please revise.".format(k))
        ahp_wts_dict[k] = ahp

    print("=====================================================================")
    print("Second Phase done - Read relative Consistency File. Multiplying Results with Weights.")
    print("=====================================================================")
    l2wts = ahp_wts_dict["L2"]
    nl2wts = l2wts.eigdf * ahp_wts_dict["L1"].eigdf.loc["Computation"]
    l2wts.eigdf = nl2wts

    # L3
    l31wts = ahp_wts_dict["L3.1"]
    nl31wts = l31wts.eigdf * ahp_wts_dict["L2"].eigdf.loc["Performance"]
    l31wts.eigdf = nl31wts

    l32wts = ahp_wts_dict["L3.2"]
    nl32wts = l32wts.eigdf * ahp_wts_dict["L2"].eigdf.loc["Compatibility"]
    l32wts.eigdf = nl32wts

    # Calculating the global priorites is done. Now plug in the values from bottom up
    # L3
    cpu_used_vals = ahp_dict["CPU Used"].eigdf*ahp_wts_dict["L3.1"].eigdf.loc["CPU"]
    mem_used_vals = ahp_dict["Mem Used"].eigdf*ahp_wts_dict["L3.1"].eigdf.loc["Memory"]
    cpu_free_vals = ahp_dict["CPU Free"].eigdf*ahp_wts_dict["L3.2"].eigdf.loc["CPU"]
    mem_free_vals = ahp_dict["Mem Free"].eigdf*ahp_wts_dict["L3.2"].eigdf.loc["Memory"]

    # L2
    faults_vals = ahp_dict["Faults"].eigdf*ahp_wts_dict["L2"].eigdf.loc["Reliability"]
    
    # L1
    pow_vals = ahp_dict["Power"].eigdf*ahp_wts_dict["L1"].eigdf.loc["Power"]
    wt_vals = ahp_dict["Weight"].eigdf*ahp_wts_dict["L1"].eigdf.loc["Weight"]
    vol_vals = ahp_dict["Size"].eigdf*ahp_wts_dict["L1"].eigdf.loc["Size"]

    fullcollist = ["Size", "Power", "Weight", "CPU Used", "Memory Used", "CPU Free", "Memory Free", "Faults"]
    defs = np.zeros((pow_vals.shape[0], len(fullcollist)))
    finaldf = pd.DataFrame(data=defs, index=ahp_dict["Power"].df.index, columns=fullcollist)
    # print(finaldf)
    finaldf.at[:,"Size"] = vol_vals["Relative Weight"]

    finaldf.at[:,"Weight"] = wt_vals["Relative Weight"]
    finaldf.at[:,"Power"] = pow_vals["Relative Weight"]
    finaldf.at[:,"CPU Used"] = cpu_used_vals["Relative Weight"]
    finaldf.at[:,"Memory Used"] = mem_used_vals["Relative Weight"]
    finaldf.at[:,"CPU Free"] = cpu_free_vals["Relative Weight"]
    finaldf.at[:,"Memory Free"] = mem_free_vals["Relative Weight"]
    finaldf.at[:,"Faults"] = faults_vals["Relative Weight"]
    finaldf["Sum"] = finaldf.sum(axis=1)
    bestsol = finaldf["Sum"].idxmax()
    print("Best Option with given Parameters is: {}".format(bestsol))
    finaldf.index = finaldf.index.droplevel("compression")
    print(finaldf)
    # Function to write to file
    written = writeToFile(finaldf, floc)
    if written:
        print("Written to output sheet into same file")
    else:
        print("Already existed. Not Written")
    print("Test Done")
    # fig, ax = plt.subplots()
    # linesofinterest = resorted_dict["CPU Load max"]["piuncompr_1_nico-pi"]
    # ax.plot(linesofinterest)
    # ax.set_ylim(0,)
    # plt.show()
    plot_node_df_dict(resorted_dict)

    print("Done")